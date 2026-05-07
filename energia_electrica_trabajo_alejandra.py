import heapq
import networkx as nx
import matplotlib.pyplot as plt
import math
import random
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


MENSAJE_CIERZO = """
╔══════════════════════════════════════════════════════════════════════════════╗
║        RED DE DISTRIBUCIÓN ELÉCTRICA – ZARAGOZA COMO NODO CENTRAL            ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                              ║
║  EL CIERZO Y LA ENERGÍA EÓLICA DE ARAGÓN                                     ║
║                                                                              ║
║  El cierzo es un viento del noroeste que recorre el Valle del Ebro con una   ║
║  intensidad y regularidad excepcionales, resultado de la orografía única de  ║
║  Aragón: los Pirineos al norte y el Sistema Ibérico al sur crean un canal    ║
║  natural que encauza y acelera el viento hasta velocidades de 100 km/h.      ║
║                                                                              ║
║  Esta característica convierte a Aragón en la segunda comunidad autónoma     ║
║  de España en potencia eólica instalada. La región genera casi el doble de   ║
║  energía eléctrica de la que consume, exportando el excedente al resto del   ║
║  país a través de las líneas de alta tensión de Red Eléctrica de España.     ║
║                                                                              ║
║  ZARAGOZA COMO CENTRO DE DISTRIBUCIÓN                                        ║
║                                                                              ║
║  Zaragoza ocupa una posición geográfica privilegiada: está a unos 300 km     ║
║  de las ciudades más importantes del país como son: Madrid, Barcelona,       ║
║  Valencia, Bilbao y San Sebastián, y conecta con las capitales aragonesas    ║
║  de Huesca y Teruel. Esto la convierte en el nodo central natural desde      ║
║  el que distribuir la energía generada en Aragón.                            ║
║                                                                              ║
║  ESTA SIMULACIÓN                                                             ║
║                                                                              ║
║  Modelamos la red eléctrica peninsular completa (47 provincias) como un      ║
║  grafo dirigido y ponderado. Cada nodo es una capital de provincia y cada    ║
║  arista representa una línea de alta tensión cuyo peso es la distancia en    ║
║  km entre provincias vecinas. El algoritmo de DIJKSTRA calcula la ruta de    ║
║  menor pérdida energética desde Zaragoza hasta cualquier punto de España.    ║
║  Además, simulamos averías aleatorias en la ruta elegida para estudiar la    ║
║  utilidad de la red y el impacto sobre el suministro eléctrico de España.    ║
║                                                                              ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

#   BASE DE DATOS: COORDENADAS DE LAS 47 PROVINCIAS PENINSULARES

coordenadas = {
    "Zaragoza":      (41.6488, -0.8891),
    "Huesca":        (42.1401, -0.4089),
    "Teruel":        (40.3456, -1.1065),
    "Barcelona":     (41.3874,  2.1686),
    "Girona":        (41.9794,  2.8214),
    "Lleida":        (41.6176,  0.6200),
    "Tarragona":     (41.1189,  1.2445),
    "Valencia":      (39.4699, -0.3763),
    "Castellon":     (39.9864, -0.0513),
    "Alicante":      (38.3452, -0.4810),
    "Madrid":        (40.4168, -3.7038),
    "Guadalajara":   (40.6334, -3.1674),
    "Cuenca":        (40.0704, -2.1374),
    "Toledo":        (39.8628, -4.0273),
    "Ciudad_Real":   (38.9848, -3.9274),
    "Albacete":      (38.9943, -1.8585),
    "Soria":         (41.7640, -2.4650),
    "Segovia":       (40.9429, -4.1088),
    "Avila":         (40.6567, -4.6976),
    "Salamanca":     (40.9701, -5.6635),
    "Zamora":        (41.5034, -5.7447),
    "Valladolid":    (41.6523, -4.7245),
    "Palencia":      (42.0097, -4.5288),
    "Leon":          (42.5987, -5.5671),
    "Burgos":        (42.3439, -3.6970),
    "Logrono":       (42.4627, -2.4449),
    "Pamplona":      (42.8125, -1.6458),
    "Bilbao":        (43.2630, -2.9350),
    "San_Sebastian": (43.3183, -1.9812),
    "Vitoria":       (42.8467, -2.6727),
    "Santander":     (43.4623, -3.8099),
    "Oviedo":        (43.3614, -5.8593),
    "Lugo":          (43.0097, -7.5560),
    "Ourense":       (42.3354, -7.8639),
    "Pontevedra":    (42.4336, -8.6488),
    "A_Coruna":      (43.3623, -8.4115),
    "Caceres":       (39.4753, -6.3724),
    "Badajoz":       (38.8794, -6.9706),
    "Huelva":        (37.2614, -6.9447),
    "Sevilla":       (37.3891, -5.9845),
    "Cadiz":         (36.5271, -6.2886),
    "Malaga":        (36.7213, -4.4214),
    "Cordoba":       (37.8882, -4.7794),
    "Jaen":          (37.7796, -3.7849),
    "Granada":       (37.1773, -3.5986),
    "Almeria":       (36.8340, -2.4637),
    "Murcia":        (37.9922, -1.1307),
}

#   VECINDADES REALES (conexiones entre provincias limítrofes)

vecindades = {
    "Zaragoza":    ["Huesca","Teruel","Lleida","Tarragona","Guadalajara",
                    "Soria","Logrono","Pamplona","San_Sebastian"],
    "Huesca":      ["Zaragoza","Lleida","Pamplona"],
    "Teruel":      ["Zaragoza","Cuenca","Valencia","Castellon","Lleida"],
    "Barcelona":   ["Girona","Lleida","Tarragona"],
    "Girona":      ["Barcelona","Lleida"],
    "Lleida":      ["Huesca","Zaragoza","Tarragona","Barcelona","Girona","Teruel"],
    "Tarragona":   ["Lleida","Barcelona","Zaragoza","Castellon"],
    "Valencia":    ["Castellon","Teruel","Cuenca","Albacete","Alicante"],
    "Castellon":   ["Tarragona","Teruel","Valencia"],
    "Alicante":    ["Valencia","Albacete","Murcia"],
    "Madrid":      ["Guadalajara","Segovia","Avila","Toledo","Cuenca"],
    "Guadalajara": ["Zaragoza","Soria","Madrid","Cuenca","Segovia"],
    "Cuenca":      ["Guadalajara","Teruel","Valencia","Albacete","Toledo","Madrid"],
    "Toledo":      ["Madrid","Cuenca","Ciudad_Real","Avila","Caceres"],
    "Ciudad_Real": ["Toledo","Cuenca","Albacete","Cordoba","Badajoz","Caceres"],
    "Albacete":    ["Cuenca","Valencia","Alicante","Murcia","Jaen","Ciudad_Real"],
    "Soria":       ["Zaragoza","Logrono","Burgos","Segovia","Guadalajara"],
    "Segovia":     ["Soria","Burgos","Valladolid","Avila","Madrid","Guadalajara"],
    "Avila":       ["Segovia","Valladolid","Salamanca","Caceres","Toledo","Madrid"],
    "Salamanca":   ["Zamora","Valladolid","Avila","Caceres"],
    "Zamora":      ["Leon","Valladolid","Salamanca","Ourense"],
    "Valladolid":  ["Palencia","Burgos","Soria","Segovia","Avila","Zamora","Salamanca"],
    "Palencia":    ["Leon","Burgos","Valladolid"],
    "Leon":        ["Oviedo","Lugo","Ourense","Zamora","Palencia","Burgos","Santander"],
    "Burgos":      ["Santander","Bilbao","Vitoria","Logrono","Soria",
                    "Valladolid","Palencia","Leon"],
    "Logrono":     ["Zaragoza","Pamplona","Vitoria","Burgos","Soria"],
    "Pamplona":    ["Huesca","Zaragoza","Logrono","San_Sebastian","Vitoria"],
    "Bilbao":      ["Santander","Burgos","Vitoria","San_Sebastian"],
    "San_Sebastian":["Pamplona","Bilbao","Zaragoza"],
    "Vitoria":     ["Bilbao","San_Sebastian","Pamplona","Logrono","Burgos"],
    "Santander":   ["Bilbao","Burgos","Leon","Oviedo"],
    "Oviedo":      ["Santander","Leon","Lugo"],
    "Lugo":        ["Oviedo","Leon","Ourense","A_Coruna","Pontevedra"],
    "Ourense":     ["Lugo","Zamora","Pontevedra","A_Coruna"],
    "Pontevedra":  ["Lugo","Ourense","A_Coruna"],
    "A_Coruna":    ["Lugo","Ourense","Pontevedra"],
    "Caceres":     ["Salamanca","Avila","Toledo","Ciudad_Real","Badajoz"],
    "Badajoz":     ["Caceres","Ciudad_Real","Cordoba","Huelva","Sevilla"],
    "Huelva":      ["Badajoz","Sevilla"],
    "Sevilla":     ["Huelva","Badajoz","Cordoba","Malaga","Cadiz"],
    "Cadiz":       ["Sevilla","Malaga"],
    "Malaga":      ["Cadiz","Sevilla","Cordoba","Granada"],
    "Cordoba":     ["Badajoz","Ciudad_Real","Albacete","Jaen","Sevilla","Malaga"],
    "Jaen":        ["Cordoba","Albacete","Granada","Ciudad_Real"],
    "Granada":     ["Jaen","Malaga","Almeria"],
    "Almeria":     ["Granada","Murcia"],
    "Murcia":      ["Albacete","Alicante","Almeria"],
}


#   PARÁMETROS ENERGÉTICOS
#   Pérdida: 4% de la energía inicial por cada 100 km recorridos.
#   Referencia orientativa para líneas de alta tensión de 400 kV.

ENERGIA_INICIAL    = 1000.0   # MWh disponibles en Zaragoza
PERDIDA_CADA_100KM = 4.0      # % de ENERGIA_INICIAL perdido por cada 100 km
COSTE_POR_KM       = 500_000  # € por km de línea de alta tensión (ref. orientativa)
ORIGEN = "Zaragoza"


def distancia_geo(c1, c2):
    """Distancia geodésica en km entre dos pares (lat, lon)."""
    R = 6371
    lat1, lon1 = c1; lat2, lon2 = c2
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon/2)**2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def calcular_energia_final(distancia_km):
    """E_final = E_inicial − (dist/100) × (P%/100) × E_inicial"""
    perdida = (distancia_km / 100.0) * (PERDIDA_CADA_100KM / 100.0) * ENERGIA_INICIAL
    return max(ENERGIA_INICIAL - perdida, 0.0)


def calcular_coste(distancia_km):
    return distancia_km * COSTE_POR_KM


#   CLASE GRAFO
class Grafo:
    def __init__(self):
        self.adyacencia = {}

    def agregar_ciudad(self, ciudad):
        if ciudad not in self.adyacencia:
            self.adyacencia[ciudad] = []

    def agregar_conexion(self, origen, destino, distancia):
        self.agregar_ciudad(origen)
        self.agregar_ciudad(destino)
        self.adyacencia[origen].append((destino, round(distancia)))

    def eliminar_conexion(self, origen, destino):
        if origen in self.adyacencia:
            self.adyacencia[origen] = [
                (v, p) for v, p in self.adyacencia[origen] if v != destino
            ]

    def obtener_vecinos(self, ciudad):
        return self.adyacencia.get(ciudad, [])

    def listar_conexiones(self):
        return [(o, d, p) for o, vs in self.adyacencia.items() for d, p in vs]


def construir_grafo():
    g = Grafo()
    for origen, vecinos in vecindades.items():
        if origen not in coordenadas:
            continue
        for destino in vecinos:
            if destino not in coordenadas:
                continue
            dist = distancia_geo(coordenadas[origen], coordenadas[destino])
            g.agregar_conexion(origen, destino, dist)
    return g


#   DIJKSTRA

def dijkstra(grafo, origen):
    """
    Calcula la ruta de menor distancia acumulada (= menor pérdida energética)
    desde 'origen' a todos los nodos alcanzables, usando un min-heap.

    Retorna:
        distancias: {ciudad: km_mínimos}
        previos:    {ciudad: nodo_anterior}
    """
    distancias = {c: float('inf') for c in grafo.adyacencia}
    distancias[origen] = 0
    previos = {c: None for c in grafo.adyacencia}
    cola = [(0, origen)]

    while cola:
        dist_actual, actual = heapq.heappop(cola)
        if dist_actual > distancias[actual]:
            continue
        for vecino, peso in grafo.obtener_vecinos(actual):
            nueva = dist_actual + peso
            if nueva < distancias[vecino]:
                distancias[vecino] = nueva
                previos[vecino] = actual
                heapq.heappush(cola, (nueva, vecino))

    return distancias, previos

def reconstruir_camino(previos, destino):
    camino, actual = [], destino
    while actual is not None:
        camino.append(actual)
        actual = previos[actual]
    camino.reverse()
    return camino


#   SIMULACIÓN

def ejecutar_simulacion(grafo):
    distancias, previos = dijkstra(grafo, ORIGEN)
    resultados = []
    for ciudad in grafo.adyacencia:
        dist   = distancias[ciudad]
        camino = reconstruir_camino(previos, ciudad)
        energia = calcular_energia_final(dist) if dist != float('inf') else 0.0
        coste   = calcular_coste(dist)         if dist != float('inf') else 0.0
        resultados.append({
            "ciudad":    ciudad,
            "distancia": dist,
            "energia":   energia,
            "coste":     coste,
            "camino":    camino,
            "saltos":    len(camino) - 1,
        })
    return resultados

def imprimir_resultados(resultados, titulo="SITUACIÓN NORMAL"):
    print(f"\n{'='*72}")
    print(f"  {titulo}")
    print(f"  Pérdida: {PERDIDA_CADA_100KM}% de {ENERGIA_INICIAL:.0f} MWh por cada 100 km")
    print(f"{'='*72}")
    print(f"  {'Provincia':18} {'Dist (km)':>10} {'Energía (MWh)':>14} {'Coste línea (€)':>18}  Ruta")
    print(f"  {'-'*70}")
    for r in sorted(resultados, key=lambda x: x["distancia"]):
        if r["distancia"] == 0:
            continue
        if r["distancia"] == float('inf'):
            print(f"  {r['ciudad']:18} {'SIN RUTA':>10} {'0.0':>14} {'-':>18}")
            continue
        ruta = " → ".join(r["camino"])
        print(f"  {r['ciudad']:18} {r['distancia']:>10.0f} "
              f"{r['energia']:>14.1f} "
              f"{r['coste']:>18,.0f}  {ruta}")
    print()

#   EXPORTACIÓN A EXCEL (openpyxl)

def exportar_excel(resultados_normal, resultados_fallo, fallo,
                   ruta_archivo="red_electrica.xlsx"):

    origen_f, destino_f, km_f = fallo
    norm_d  = {r["ciudad"]: r for r in resultados_normal}
    fallo_d = {r["ciudad"]: r for r in resultados_fallo}

    wb = Workbook()

    COLOR_HEADER    = "1A3A5C"   # azul marino
    COLOR_SUBHEADER = "2E6DA4"   # azul medio
    COLOR_NORMAL    = "D6E4F0"   # azul muy claro
    COLOR_FALLO     = "FADBD8"   # rojo muy claro
    COLOR_AVISO     = "F9E79F"   # amarillo claro
    COLOR_SIN_SUM   = "922B21"   # rojo oscuro
    COLOR_MEJORA    = "1E8449"   # verde oscuro
    COLOR_BLANCO    = "FFFFFF"

    borde = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    def estilo_celda(cell, bold=False, color_fondo=None, color_fuente="000000",
                     alineacion="left", tamaño=10, borde_activo=True):
        cell.font = Font(name="Arial", bold=bold, color=color_fuente, size=tamaño)
        if color_fondo:
            cell.fill = PatternFill("solid", start_color=color_fondo)
        cell.alignment = Alignment(horizontal=alineacion, vertical="center",
                                   wrap_text=True)
        if borde_activo:
            cell.border = borde

    ws_info = wb.active
    ws_info.title = "Parámetros"
    ws_info.column_dimensions["A"].width = 35
    ws_info.column_dimensions["B"].width = 45

    filas_info = [
        ("RED DE DISTRIBUCIÓN ELÉCTRICA", None, True, COLOR_HEADER, "FFFFFF", 14),
        ("Simulación con algoritmo de Dijkstra", None, False, "2E6DA4", "FFFFFF", 11),
        (None, None, False, None, None, None),
        ("PARÁMETROS DE LA SIMULACIÓN", None, True, COLOR_SUBHEADER, "FFFFFF", 11),
        ("Nodo origen", ORIGEN, False, COLOR_NORMAL, "000000", 10),
        ("Energía inicial", f"{ENERGIA_INICIAL:.0f} MWh", False, COLOR_NORMAL, "000000", 10),
        ("Pérdida energética", f"{PERDIDA_CADA_100KM}% por cada 100 km", False, COLOR_NORMAL, "000000", 10),
        ("Coste estimado por km", f"{COSTE_POR_KM:,.0f} €/km (referencia 400 kV)", False, COLOR_NORMAL, "000000", 10),
        (None, None, False, None, None, None),
        ("AVERÍA SIMULADA", None, True, "7B241C", "FFFFFF", 11),
        ("Tramo cortado", f"{origen_f} → {destino_f}", False, COLOR_FALLO, "000000", 10),
        ("Distancia del tramo", f"{km_f} km", False, COLOR_FALLO, "000000", 10),
        (None, None, False, None, None, None),
        ("DESCRIPCIÓN", None, True, COLOR_SUBHEADER, "FFFFFF", 11),
        ("Modelo",
         "Grafo dirigido y ponderado. Cada nodo es una capital de provincia "
         "y cada arista una línea de alta tensión.", False, COLOR_BLANCO, "000000", 10),
        ("Algoritmo",
         "Dijkstra: calcula la ruta de menor distancia acumulada "
         "(equivalente a menor pérdida energética).", False, COLOR_BLANCO, "000000", 10),
        ("Fallo",
         "Se elimina aleatoriamente un tramo de la ruta óptima y se recalcula "
         "Dijkstra para encontrar la mejor alternativa.", False, COLOR_BLANCO, "000000", 10),
    ]

    for fila in filas_info:
        etiqueta, valor, bold, c_fondo, c_fuente, tam = fila
        if etiqueta is None:
            ws_info.append([])
            continue
        ws_info.append([etiqueta, valor or ""])
        row = ws_info.max_row
        for col in [1, 2]:
            cell = ws_info.cell(row=row, column=col)
            estilo_celda(cell, bold=bold,
                         color_fondo=c_fondo if c_fondo else COLOR_BLANCO,
                         color_fuente=c_fuente or "000000",
                         tamaño=tam or 10,
                         borde_activo=bool(c_fondo))
        if valor is None:
            ws_info.merge_cells(f"A{row}:B{row}")
            ws_info.cell(row=row, column=1).alignment = Alignment(
                horizontal="center", vertical="center")

    ws_info.row_dimensions[1].height = 28
    ws_info.row_dimensions[2].height = 20

    ws = wb.create_sheet("Resultados")

    cabeceras = [
        "Provincia", "Com. Autónoma",
        "Dist. normal (km)", "Energía normal (MWh)", "Pérdida normal (MWh)",
        "Coste línea (€)", "Saltos", "Ruta óptima normal",
        "Dist. avería (km)", "Energía avería (MWh)", "Pérdida avería (MWh)",
        "Saltos avería", "Ruta alternativa",
        "Δ Distancia (km)", "Δ Energía (MWh)", "Estado",
    ]

    anchos = [18, 20, 18, 20, 18, 18, 8, 55, 18, 20, 18, 12, 55, 16, 16, 18]
    for i, (cab, ancho) in enumerate(zip(cabeceras, anchos), 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Título
    ws.merge_cells("A1:P1")
    titulo = ws["A1"]
    titulo.value = f"Red eléctrica peninsular – Resultados Dijkstra | Avería: {origen_f} → {destino_f}"
    estilo_celda(titulo, bold=True, color_fondo=COLOR_HEADER,
                 color_fuente="FFFFFF", alineacion="center", tamaño=12)
    ws.row_dimensions[1].height = 22

    # Cabecera de grupos
    ws.merge_cells("A2:B2");  ws["A2"].value = "PROVINCIA"
    ws.merge_cells("C2:H2");  ws["C2"].value = "SITUACIÓN NORMAL"
    ws.merge_cells("I2:M2");  ws["I2"].value = "TRAS AVERÍA"
    ws.merge_cells("N2:P2");  ws["N2"].value = "COMPARATIVA"
    for celda, color in [("A2", COLOR_HEADER), ("C2", "1A5276"),
                          ("I2", "7B241C"),    ("N2", "145A32")]:
        estilo_celda(ws[celda], bold=True, color_fondo=color,
                     color_fuente="FFFFFF", alineacion="center", tamaño=10)
    ws.row_dimensions[2].height = 18

    # Cabecera de columnas
    for i, cab in enumerate(cabeceras, 1):
        cell = ws.cell(row=3, column=i, value=cab)
        color = (COLOR_NORMAL if i <= 8
                 else COLOR_FALLO if i <= 13
                 else "D5F5E3")
        estilo_celda(cell, bold=True, color_fondo=color,
                     alineacion="center", tamaño=9)
    ws.row_dimensions[3].height = 32

    # Datos
    for ciudad in sorted(norm_d.keys()):
        if ciudad == ORIGEN:
            continue
        rn = norm_d[ciudad]
        rf = fallo_d.get(ciudad, rn)

        dist_n, dist_f = rn["distancia"], rf["distancia"]
        en_n,   en_f   = rn["energia"],   rf["energia"]
        sin_sum = dist_f == float('inf')

        delta_km = (dist_f - dist_n) if not sin_sum else None
        delta_en = (en_n - en_f)     if not sin_sum else None

        if sin_sum:
            estado = "SIN SUMINISTRO"
        elif delta_km and delta_km > 0:
            estado = "Ruta alternativa"
        else:
            estado = "Sin cambios"

        ccaa = _ciudad_a_ccaa(ciudad)

        fila = [
            ciudad, ccaa,
            int(dist_n) if dist_n != float('inf') else "SIN RUTA",
            round(en_n, 1),
            round(ENERGIA_INICIAL - en_n, 1),
            int(rn["coste"]),
            rn["saltos"],
            " → ".join(rn["camino"]),
            int(dist_f) if not sin_sum else "SIN RUTA",
            round(en_f, 1) if not sin_sum else 0,
            round(ENERGIA_INICIAL - en_f, 1) if not sin_sum else ENERGIA_INICIAL,
            rf["saltos"] if not sin_sum else "—",
            " → ".join(rf["camino"]) if not sin_sum else "Sin ruta",
            int(delta_km) if delta_km is not None else "—",
            round(delta_en, 1) if delta_en is not None else "—",
            estado,
        ]

        row_idx = ws.max_row + 1
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

            # Color de fondo según zona y estado
            if sin_sum and col_idx >= 9:
                estilo_celda(cell, color_fondo="FADBD8", color_fuente="7B241C")
            elif col_idx <= 2:
                estilo_celda(cell, bold=(col_idx == 1),
                             color_fondo="EBF5FB" if row_idx % 2 == 0 else COLOR_BLANCO)
            elif col_idx <= 8:
                estilo_celda(cell, color_fondo="EBF5FB" if row_idx % 2 == 0 else COLOR_BLANCO)
            elif col_idx <= 13:
                c = "FADBD8" if sin_sum else ("FEF9E7" if row_idx % 2 == 0 else COLOR_BLANCO)
                estilo_celda(cell, color_fondo=c)
            else:
                if estado == "SIN SUMINISTRO":
                    estilo_celda(cell, bold=True, color_fondo="E74C3C", color_fuente="FFFFFF")
                elif estado == "Ruta alternativa":
                    estilo_celda(cell, bold=True, color_fondo=COLOR_AVISO)
                else:
                    estilo_celda(cell, color_fondo="D5F5E3")

            # Alineación numérica
            if col_idx in (3, 4, 5, 6, 7, 9, 10, 11, 12, 14, 15):
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row_idx].height = 15

    # Fila de totales / resumen
    row_tot = ws.max_row + 2
    ws.cell(row=row_tot, column=1, value="RESUMEN").font = Font(bold=True, name="Arial")
    n_sin  = sum(1 for r in resultados_fallo if r["distancia"] == float('inf')
                 and r["ciudad"] != ORIGEN)
    n_alt  = sum(1 for r in resultados_fallo
                 if r["ciudad"] != ORIGEN
                 and r["distancia"] != float('inf')
                 and r["distancia"] > norm_d.get(r["ciudad"], {}).get("distancia", 0))
    resumen = [
        (f"Provincias sin suministro tras avería: {n_sin}", "E74C3C", "FFFFFF"),
        (f"Provincias con ruta alternativa: {n_alt}",       "F39C12", "FFFFFF"),
        (f"Provincias sin cambios: {46 - n_sin - n_alt}",   "27AE60", "FFFFFF"),
    ]
    for i, (texto, c_fondo, c_fuente) in enumerate(resumen):
        cell = ws.cell(row=row_tot + i, column=1, value=texto)
        ws.merge_cells(f"A{row_tot+i}:P{row_tot+i}")
        estilo_celda(cell, bold=True, color_fondo=c_fondo,
                     color_fuente=c_fuente, alineacion="center")
        ws.row_dimensions[row_tot + i].height = 16

    # Inmovilizar cabecera
    ws.freeze_panes = "A4"

    wb.save(ruta_archivo)
    print(f" Excel exportado: {os.path.abspath(ruta_archivo)}\n")


#   MODO FALLO ALEATORIO

def simular_fallo_aleatorio(grafo):
    aristas = grafo.listar_conexiones()
    fallo   = random.choice(aristas)
    grafo.eliminar_conexion(fallo[0], fallo[1])
    return fallo

def analizar_impacto_fallo(res_normal, res_fallo, fallo):
    origen_f, destino_f, km_f = fallo
    print(f"\n{'='*72}")
    print(f"  ⚡ FALLO ALEATORIO SIMULADO")
    print(f"  Línea cortada: {origen_f} → {destino_f}  ({km_f} km)")
    print(f"{'='*72}")

    norm_d  = {r["ciudad"]: r for r in res_normal}
    fallo_d = {r["ciudad"]: r for r in res_fallo}

    sin_sum = []
    alternativas = []

    for ciudad, rf in fallo_d.items():
        if ciudad == ORIGEN:
            continue
        rn = norm_d[ciudad]
        if rf["distancia"] == float('inf'):
            sin_sum.append(ciudad)
        elif rf["distancia"] > rn["distancia"]:
            alternativas.append((ciudad, rf,
                                 rf["distancia"] - rn["distancia"],
                                 rn["energia"]   - rf["energia"]))

    if sin_sum:
        print(f"\n  ⚠  Provincias SIN SUMINISTRO:")
        for c in sin_sum:
            print(f"       • {c}")
    else:
        print(f"\n  ✓  Todas las provincias conservan ruta alternativa.")

    if alternativas:
        print(f"\n  Provincias con ruta alternativa (mayor pérdida):")
        print(f"  {'Provincia':18} {'Dist. nueva':>11} {'Energía':>10} {'Δ km':>8} {'Δ pérdida':>11}")
        print(f"  {'-'*62}")
        for ciudad, rf, aum, perd in sorted(alternativas, key=lambda x: -x[2]):
            ruta = " → ".join(rf["camino"])
            print(f"  {ciudad:18} {rf['distancia']:>8.0f} km "
                  f"{rf['energia']:>8.1f} MWh "
                  f"{aum:>+7.0f} km  {perd:>+9.1f} MWh")
            print(f"    Ruta alternativa: {ruta}")
    else:
        print(f"\n  Ninguna provincia ve empeorada su ruta.")
    print()


#   COLORES POR COMUNIDAD AUTÓNOMA

CCAA = {
    "Aragón":               ["Zaragoza", "Huesca", "Teruel"],
    "Cataluña":             ["Barcelona", "Girona", "Lleida", "Tarragona"],
    "C. Valenciana":        ["Valencia", "Castellon", "Alicante"],
    "Madrid":               ["Madrid"],
    "Castilla-La Mancha":   ["Guadalajara", "Cuenca", "Toledo", "Ciudad_Real", "Albacete"],
    "Castilla y León":      ["Soria", "Segovia", "Avila", "Salamanca",
                             "Zamora", "Valladolid", "Palencia", "Leon", "Burgos"],
    "La Rioja":             ["Logrono"],
    "Navarra":              ["Pamplona"],
    "País Vasco":           ["Bilbao", "San_Sebastian", "Vitoria"],
    "Cantabria":            ["Santander"],
    "Asturias":             ["Oviedo"],
    "Galicia":              ["Lugo", "Ourense", "Pontevedra", "A_Coruna"],
    "Extremadura":          ["Caceres", "Badajoz"],
    "Andalucía":            ["Huelva", "Sevilla", "Cadiz", "Malaga",
                             "Cordoba", "Jaen", "Granada", "Almeria"],
    "Murcia":               ["Murcia"],
}

# Paleta de colores distinguibles para cada CCAA
COLORES_CCAA = {
    "Aragón":               "#D94F3D",   # rojo ladrillo
    "Cataluña":             "#E8A838",   # ámbar dorado
    "C. Valenciana":        "#4CAF50",   # verde esmeralda
    "Madrid":               "#7B68EE",   # azul medio lavanda
    "Castilla-La Mancha":   "#26A69A",   # verde azulado
    "Castilla y León":      "#42A5F5",   # azul cielo
    "La Rioja":             "#EC407A",   # rosa frambuesa
    "Navarra":              "#AB47BC",   # violeta medio
    "País Vasco":           "#26C6DA",   # cian turquesa
    "Cantabria":            "#8D6E63",   # marrón cálido
    "Asturias":             "#FF7043",   # naranja coral
    "Galicia":              "#5C6BC0",   # índigo
    "Extremadura":          "#66BB6A",   # verde lima
    "Andalucía":            "#EF5350",   # rojo vivo
    "Murcia":               "#FFA726",   # naranja ámbar
}

# Mapa inverso: ciudad → color de su CCAA
def _color_ccaa(ciudad):
    for ccaa, ciudades in CCAA.items():
        if ciudad in ciudades:
            return COLORES_CCAA[ccaa]
    return "#aaaaaa"

def _ciudad_a_ccaa(ciudad):
    for ccaa, ciudades in CCAA.items():
        if ciudad in ciudades:
            return ccaa
    return "Desconocida"


#   VISUALIZACIÓN

def _pos():
    return {c: (coord[1], coord[0]) for c, coord in coordenadas.items()}

def _nx(grafo):
    G = nx.DiGraph()
    for o, vs in grafo.adyacencia.items():
        for d, p in vs:
            G.add_edge(o, d, weight=p)
    return G

def _leyenda_ccaa(ax):
    """Añade una leyenda de CCAA con sus colores al eje dado."""
    from matplotlib.patches import Patch
    handles = [Patch(facecolor=color, edgecolor="white", label=ccaa)
               for ccaa, color in COLORES_CCAA.items()]
    ax.legend(handles=handles, title="Comunidad Autónoma",
              loc="lower left", fontsize=6, title_fontsize=7,
              framealpha=0.85, ncol=2)


def dibujar_red_normal(grafo, resultados):
    """
    Mapa general de la red con nodos coloreados por CCAA.
    Muestra los MWh que llegan a cada provincia.
    """
    G, pos = _nx(grafo), _pos()
    e_dict = {r["ciudad"]: r["energia"] for r in resultados}

    fig, ax = plt.subplots(figsize=(17, 11))

    # Dibujar nodos agrupados por CCAA
    for ccaa, ciudades in CCAA.items():
        nodos = [c for c in ciudades if c in G.nodes()]
        if not nodos:
            continue
        nx.draw_networkx_nodes(G, pos, nodelist=nodos,
                               node_size=550,
                               node_color=COLORES_CCAA[ccaa],
                               edgecolors="white", linewidths=1.2,
                               ax=ax)

    # Nodo origen con borde dorado especial
    nx.draw_networkx_nodes(G, pos, nodelist=[ORIGEN],
                           node_size=750,
                           node_color=COLORES_CCAA["Aragón"],
                           edgecolors="gold", linewidths=3, ax=ax)

    # Etiquetas: nombre + MWh
    for n in G.nodes():
        mwh = e_dict.get(n, ENERGIA_INICIAL)
        label = f"{n}\n{mwh:.0f} MWh" if n != ORIGEN else f"{n}\n⚡ {ENERGIA_INICIAL:.0f} MWh"
        nx.draw_networkx_labels(G, pos, labels={n: label},
                                font_size=5.5, font_weight="bold",
                                font_color="white", ax=ax)

    nx.draw_networkx_edges(G, pos, arrowstyle="->", arrowsize=10,
                           edge_color="#666666", width=0.7, ax=ax)

    edge_labels = {(u, v): f"{d['weight']} km" for u, v, d in G.edges(data=True)}
    nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels,
                                 font_size=4, label_pos=0.35,
                                 bbox=dict(boxstyle="round,pad=0.1",
                                           fc="white", alpha=0.65), ax=ax)

    _leyenda_ccaa(ax)
    ax.set_title(
        f"Red eléctrica peninsular – nodos coloreados por Comunidad Autónoma\n"
        f"Pérdida: {PERDIDA_CADA_100KM}% de {ENERGIA_INICIAL:.0f} MWh por cada 100 km  "
        f"| ⭐ = nodo origen",
        fontsize=11, fontweight="bold"
    )
    ax.set_axis_off()
    plt.tight_layout()
    plt.show()

def dibujar_ruta_destino(grafo, resultados, destino):
    """
    Resalta la ruta óptima hacia el destino.
    - Nodos fuera del camino: color CCAA tenue
    - Nodos del camino intermedio: color CCAA saturado + borde naranja grueso
    - Origen: verde con borde dorado
    - Destino: rojo con borde blanco
    - Aristas del camino: flecha naranja gruesa
    """
    G, pos = _nx(grafo), _pos()
    dato = next((r for r in resultados if r["ciudad"] == destino), None)
    if dato is None or dato["distancia"] == float('inf'):
        print(f"  No hay ruta disponible hasta {destino}.")
        return

    camino        = dato["camino"]
    aristas_camino = list(zip(camino[:-1], camino[1:]))
    nodos_camino   = set(camino)
    e_dict = {r["ciudad"]: r["energia"] for r in resultados}

    fig, ax = plt.subplots(figsize=(17, 11))

    # Nodos fuera del camino: color CCAA con alpha reducido
    for ccaa, ciudades in CCAA.items():
        nodos_fuera = [c for c in ciudades
                       if c in G.nodes() and c not in nodos_camino]
        if nodos_fuera:
            nx.draw_networkx_nodes(G, pos, nodelist=nodos_fuera,
                                   node_size=350,
                                   node_color=COLORES_CCAA[ccaa],
                                   edgecolors="white", linewidths=0.8,
                                   alpha=0.35, ax=ax)

    # Nodos intermedios del camino: color CCAA + borde naranja
    intermedios = [n for n in nodos_camino if n not in (ORIGEN, destino)]
    for n in intermedios:
        nx.draw_networkx_nodes(G, pos, nodelist=[n],
                               node_size=650,
                               node_color=_color_ccaa(n),
                               edgecolors="#ff8c00", linewidths=3.5,
                               ax=ax)

    # Origen
    nx.draw_networkx_nodes(G, pos, nodelist=[ORIGEN],
                           node_size=800, node_color="#2ecc71",
                           edgecolors="gold", linewidths=3.5, ax=ax)
    # Destino
    nx.draw_networkx_nodes(G, pos, nodelist=[destino],
                           node_size=800, node_color="#e74c3c",
                           edgecolors="white", linewidths=3.5, ax=ax)

    # Etiquetas
    for n in G.nodes():
        en_camino = n in nodos_camino
        if en_camino:
            label = f"{n}\n{e_dict.get(n, ENERGIA_INICIAL):.0f} MWh"
            nx.draw_networkx_labels(G, pos, labels={n: label},
                                    font_size=7.5, font_weight="bold",
                                    font_color="white", ax=ax)
        else:
            nx.draw_networkx_labels(G, pos, labels={n: n},
                                    font_size=5, font_color="#999999", ax=ax)

    # Aristas normales (tenues)
    nx.draw_networkx_edges(G, pos, arrowstyle="->", arrowsize=8,
                           edge_color="#dddddd", width=0.5, ax=ax)
    # Aristas del camino (naranjas y gruesas)
    nx.draw_networkx_edges(G, pos, edgelist=aristas_camino,
                           edge_color="#ff6b00", width=3.5,
                           arrowstyle="->", arrowsize=20, ax=ax)

    # Km solo en aristas del camino
    ek = {(u, v): f"{G[u][v]['weight']} km" for u, v in aristas_camino}
    nx.draw_networkx_edge_labels(G, pos, edge_labels=ek,
                                 font_size=8, font_color="#7f3300",
                                 label_pos=0.4,
                                 bbox=dict(boxstyle="round,pad=0.25",
                                           fc="white", alpha=0.9), ax=ax)

    _leyenda_ccaa(ax)
    ax.set_title(
        f"Ruta óptima: {ORIGEN} → {destino}  "
        f"({_ciudad_a_ccaa(destino)})\n"
        f"{dato['distancia']:.0f} km  ·  {dato['energia']:.1f} MWh recibidos  ·  "
        f"Pérdida: {ENERGIA_INICIAL - dato['energia']:.1f} MWh  ·  "
        f"{dato['saltos']} salto(s)  ·  Coste estimado: {dato['coste']:,.0f} €",
        fontsize=10, fontweight="bold"
    )
    ax.set_axis_off()
    plt.tight_layout()
    plt.show()


def dibujar_comparacion_destino(g_orig, g_averia,
                                 dato_normal, dato_averia,
                                 fallo, destino):
    """
    Dos mapas lado a lado:
      izquierda → ruta óptima (nodos CCAA + camino en naranja)
      derecha   → ruta alternativa tras avería (tramo cortado en rojo discontinuo)
    """
    origen_f, destino_f, km_f = fallo
    pos    = _pos()
    G_orig = _nx(g_orig)
    G_av   = _nx(g_averia)

    def _pintar(ax, G, G_ref, camino, titulo,
                color_arista="#ff6b00", arista_cortada=None, sin_destino=False):

        aristas_c = list(zip(camino[:-1], camino[1:])) if camino else []
        nodos_c   = set(camino)

        # Todos los nodos: color CCAA tenue
        for ccaa, ciudades in CCAA.items():
            fuera = [c for c in ciudades if c in G_ref.nodes() and c not in nodos_c]
            if fuera:
                nx.draw_networkx_nodes(G_ref, pos, nodelist=fuera,
                                       node_size=300,
                                       node_color=COLORES_CCAA[ccaa],
                                       edgecolors="white", linewidths=0.7,
                                       alpha=0.3, ax=ax)

        # Intermedios del camino
        intermedios = [n for n in nodos_c if n not in (ORIGEN, destino)]
        for n in intermedios:
            nx.draw_networkx_nodes(G_ref, pos, nodelist=[n],
                                   node_size=600,
                                   node_color=_color_ccaa(n),
                                   edgecolors="#ff8c00", linewidths=3,
                                   ax=ax)

        # Origen
        nx.draw_networkx_nodes(G_ref, pos, nodelist=[ORIGEN],
                               node_size=700, node_color="#2ecc71",
                               edgecolors="gold", linewidths=3, ax=ax)

        # Destino
        color_dest = "#555555" if sin_destino else "#e74c3c"
        if destino in G_ref.nodes():
            nx.draw_networkx_nodes(G_ref, pos, nodelist=[destino],
                                   node_size=700, node_color=color_dest,
                                   edgecolors="white", linewidths=3, ax=ax)

        # Etiquetas
        for n in G_ref.nodes():
            en_c = n in nodos_c
            if en_c:
                e_val = (dato_averia["energia"] if dato_averia and n == destino
                         else dato_normal["energia"] if n == destino
                         else "")
                label = f"{n}\n{e_val:.0f} MWh" if e_val != "" else n
                nx.draw_networkx_labels(G_ref, pos, labels={n: label},
                                        font_size=7, font_weight="bold",
                                        font_color="white", ax=ax)
            else:
                nx.draw_networkx_labels(G_ref, pos, labels={n: n},
                                        font_size=4.5, font_color="#aaaaaa",
                                        ax=ax)

        # Aristas
        nx.draw_networkx_edges(G_ref, pos, arrowstyle="->", arrowsize=8,
                               edge_color="#dddddd", width=0.5, ax=ax)
        if aristas_c:
            nx.draw_networkx_edges(G_ref, pos, edgelist=aristas_c,
                                   edge_color=color_arista, width=3,
                                   arrowstyle="->", arrowsize=18, ax=ax)
            ek = {(u, v): f"{G_ref[u][v]['weight']} km"
                  for u, v in aristas_c if G_ref.has_edge(u, v)}
            nx.draw_networkx_edge_labels(G_ref, pos, edge_labels=ek,
                                         font_size=7, label_pos=0.4,
                                         bbox=dict(boxstyle="round,pad=0.2",
                                                   fc="white", alpha=0.88),
                                         ax=ax)

        # Tramo cortado en rojo discontinuo
        if arista_cortada and arista_cortada in G_orig.edges():
            nx.draw_networkx_edges(G_orig, pos, edgelist=[arista_cortada],
                                   edge_color="crimson", width=3.5,
                                   style="dashed",
                                   arrowstyle="->", arrowsize=20, ax=ax)
            # Etiqueta ✂ sobre el tramo cortado
            mid_x = (pos[arista_cortada[0]][0] + pos[arista_cortada[1]][0]) / 2
            mid_y = (pos[arista_cortada[0]][1] + pos[arista_cortada[1]][1]) / 2
            ax.text(mid_x, mid_y, "✂ AVERÍA", fontsize=8, color="crimson",
                    fontweight="bold", ha="center", va="center",
                    bbox=dict(boxstyle="round,pad=0.3", fc="white",
                              ec="crimson", alpha=0.9))

        _leyenda_ccaa(ax)
        ax.set_title(titulo, fontsize=9, fontweight="bold")
        ax.set_axis_off()

    fig, axes = plt.subplots(1, 2, figsize=(28, 12))

    # Mapa izquierdo: ruta óptima
    _pintar(axes[0], G_orig, G_orig,
            dato_normal["camino"],
            f"✅  RUTA ÓPTIMA: {ORIGEN} → {destino}\n"
            f"{dato_normal['distancia']:.0f} km  ·  "
            f"{dato_normal['energia']:.1f} MWh recibidos  ·  "
            f"Pérdida: {ENERGIA_INICIAL - dato_normal['energia']:.1f} MWh",
            color_arista="#27ae60")

    # Mapa derecho: tras avería
    if dato_averia is None or dato_averia["distancia"] == float('inf'):
        _pintar(axes[1], G_av, G_orig, [],
                f"⚠️  AVERÍA: {origen_f} → {destino_f} ({km_f} km)\n"
                f"{destino} queda SIN SUMINISTRO — no hay ruta alternativa",
                color_arista="red",
                arista_cortada=(origen_f, destino_f),
                sin_destino=True)
    else:
        aum  = dato_averia["distancia"] - dato_normal["distancia"]
        perd = dato_normal["energia"]   - dato_averia["energia"]
        _pintar(axes[1], G_av, G_orig,
                dato_averia["camino"],
                f"🔀  AVERÍA {origen_f} → {destino_f} ({km_f} km)  →  Ruta alternativa\n"
                f"{dato_averia['distancia']:.0f} km ({aum:+.0f} km)  ·  "
                f"{dato_averia['energia']:.1f} MWh ({-perd:+.1f} MWh)  ·  "
                f"Pérdida extra: {perd:.1f} MWh",
                color_arista="#e74c3c",
                arista_cortada=(origen_f, destino_f))

    plt.suptitle(
        f"Distribución eléctrica: {ORIGEN} → {destino}  |  "
        f"Nodos coloreados por Comunidad Autónoma  |  "
        f"Avería simulada: {origen_f} → {destino_f}",
        fontsize=12, fontweight="bold"
    )
    plt.tight_layout()
    plt.show()


#   ANÁLISIS COMBINADO: RUTA ÓPTIMA + FALLO EN ESA RUTA + ALTERNATIVA

def analizar_destino_con_fallo(grafo_original, destino):
    """
    Flujo principal centrado en un destino concreto:
      1. Calcula la ruta óptima hasta el destino (Dijkstra normal)
      2. Elige aleatoriamente un tramo de ESA ruta y lo rompe
      3. Recalcula Dijkstra sin ese tramo
      4. Muestra la comparación: ruta ideal vs. ruta alternativa tras avería
      5. Exporta el CSV con ambas situaciones
    """
    # Paso 1: ruta óptima sin averías
    res_normal = ejecutar_simulacion(grafo_original)
    dato_normal = next((r for r in res_normal if r["ciudad"] == destino), None)

    if dato_normal is None or dato_normal["distancia"] == float('inf'):
        print(f"\n  No existe ruta desde {ORIGEN} hasta {destino}.")
        return

    camino_optimo = dato_normal["camino"]
    tramos_ruta   = list(zip(camino_optimo[:-1], camino_optimo[1:]))

    print(f"\n{'='*68}")
    print(f"  RUTA ÓPTIMA SIN AVERÍAS: {ORIGEN} → {destino}")
    print(f"{'='*68}")
    print(f"  Recorrido:          {' → '.join(camino_optimo)}")
    print(f"  Distancia total:    {dato_normal['distancia']:.0f} km")
    print(f"  Energía que llega:  {dato_normal['energia']:.1f} MWh")
    print(f"  Pérdida energética: {ENERGIA_INICIAL - dato_normal['energia']:.1f} MWh")
    print(f"  Saltos:             {dato_normal['saltos']}")

    # Paso 2: fallo aleatorio en un tramo de la ruta óptima 
    tramo_fallo = random.choice(tramos_ruta)
    origen_f, destino_f = tramo_fallo
    km_fallo = next(p for v, p in grafo_original.obtener_vecinos(origen_f)
                    if v == destino_f)

    print(f"\n{'='*68}")
    print(f"  ⚡ AVERÍA SIMULADA EN LA RUTA")
    print(f"  Tramo cortado: {origen_f} → {destino_f}  ({km_fallo} km)")
    print(f"  Este tramo forma parte del recorrido óptimo.")
    print(f"{'='*68}")

    # Paso 3: recalcular con la línea rota 
    grafo_averia = construir_grafo()
    grafo_averia.eliminar_conexion(origen_f, destino_f)
    res_averia   = ejecutar_simulacion(grafo_averia)
    dato_averia  = next((r for r in res_averia if r["ciudad"] == destino), None)

    fallo = (origen_f, destino_f, km_fallo)

    # Paso 4: mostrar impacto 
    print(f"\n  RESULTADO TRAS LA AVERÍA:")
    if dato_averia is None or dato_averia["distancia"] == float('inf'):
        print(f"\n  ✗  {destino} queda SIN SUMINISTRO.")
        print(f"     No existe ninguna ruta alternativa con la línea cortada.")
    else:
        aumento_km   = dato_averia["distancia"] - dato_normal["distancia"]
        perdida_mwh  = dato_normal["energia"]   - dato_averia["energia"]
        print(f"\n  Ruta alternativa:   {' → '.join(dato_averia['camino'])}")
        print(f"  Distancia nueva:    {dato_averia['distancia']:.0f} km  "
              f"({aumento_km:+.0f} km respecto a la ruta óptima)")
        print(f"  Energía que llega:  {dato_averia['energia']:.1f} MWh  "
              f"({perdida_mwh:+.1f} MWh de pérdida adicional)")
        print(f"  Pérdida total:      {ENERGIA_INICIAL - dato_averia['energia']:.1f} MWh")
        print(f"  Saltos:             {dato_averia['saltos']}")

    # Paso 5: visualizaciones
    dibujar_ruta_destino(grafo_original, res_normal, destino)
    dibujar_comparacion_destino(grafo_original, grafo_averia,
                                 dato_normal, dato_averia,
                                 fallo, destino)

    return dato_normal, dato_averia, fallo


def dibujar_comparacion_destino(g_orig, g_averia,
                                 dato_normal, dato_averia,
                                 fallo, destino):
    """
    Dos mapas lado a lado mostrando únicamente las rutas relevantes:
      izquierda → ruta óptima normal
      derecha   → ruta alternativa tras la avería (o sin suministro)
    """
    origen_f, destino_f, km_f = fallo
    pos = _pos()
    G_orig   = _nx(g_orig)
    G_averia = _nx(g_averia)

    def _resaltar_camino(ax, G, G_ref, camino, titulo,
                         color_camino="red", arista_cortada=None):
        aristas_c  = list(zip(camino[:-1], camino[1:]))
        nodos_c    = set(camino)

        # Todos los nodos en gris
        nx.draw_networkx_nodes(G_ref, pos,
                               nodelist=[n for n in G_ref.nodes() if n not in nodos_c],
                               node_size=250, node_color="#dddddd", ax=ax)
        # Nodos del camino
        nx.draw_networkx_nodes(G_ref, pos,
                               nodelist=[n for n in nodos_c
                                         if n not in (ORIGEN, destino)],
                               node_size=500, node_color="#ff8c00", ax=ax)
        nx.draw_networkx_nodes(G_ref, pos, nodelist=[ORIGEN],
                               node_size=600, node_color="green", ax=ax)

        if destino in G_ref.nodes():
            color_dest = "black" if dato_averia is None else "red"
            nx.draw_networkx_nodes(G_ref, pos, nodelist=[destino],
                                   node_size=600, node_color=color_dest, ax=ax)

        # Etiquetas
        for n in G_ref.nodes():
            en_camino = n in nodos_c
            label = f"{n}\n{dato_normal['energia']:.0f} MWh" if n == destino and dato_averia is None else n
            if n == destino and dato_averia is not None:
                label = (f"{n}\n{dato_averia['energia']:.0f} MWh"
                         if ax.get_title().startswith("AVERÍA") or "alternativa" in ax.get_title().lower()
                         else f"{n}\n{dato_normal['energia']:.0f} MWh")
            nx.draw_networkx_labels(G_ref, pos, labels={n: label},
                                    font_size=5.5 if not en_camino else 7,
                                    font_weight="bold" if en_camino else "normal",
                                    ax=ax)

        # Todas las aristas en gris
        nx.draw_networkx_edges(G_ref, pos, arrowstyle="->", arrowsize=8,
                               edge_color="#dddddd", width=0.5, ax=ax)
        # Aristas del camino resaltadas
        if aristas_c:
            nx.draw_networkx_edges(G_ref, pos, edgelist=aristas_c,
                                   edge_color=color_camino, width=2.5,
                                   arrowstyle="->", arrowsize=16, ax=ax)

        # Tramo cortado en rojo discontinuo (solo en el mapa de avería)
        if arista_cortada and arista_cortada in G_orig.edges():
            nx.draw_networkx_edges(G_orig, pos, edgelist=[arista_cortada],
                                   edge_color="crimson", width=3, style="dashed",
                                   arrowstyle="->", arrowsize=18, ax=ax)

        # Km en las aristas del camino
        if aristas_c:
            ek = {(u, v): f"{G_ref[u][v]['weight']} km"
                  for u, v in aristas_c if G_ref.has_edge(u, v)}
            nx.draw_networkx_edge_labels(G_ref, pos, edge_labels=ek,
                                         font_size=7, font_color="darkred",
                                         label_pos=0.4,
                                         bbox=dict(boxstyle="round,pad=0.2",
                                                   fc="white", alpha=0.85),
                                         ax=ax)
        ax.set_title(titulo, fontsize=9, fontweight="bold")
        ax.set_axis_off()

    fig, axes = plt.subplots(1, 2, figsize=(26, 11))

    # Mapa izquierdo: ruta óptima normal
    _resaltar_camino(
        axes[0], G_orig, G_orig,
        dato_normal["camino"],
        f"RUTA ÓPTIMA: {ORIGEN} → {destino}\n"
        f"{dato_normal['distancia']:.0f} km  ·  {dato_normal['energia']:.1f} MWh recibidos",
        color_camino="green"
    )

    # Mapa derecho: ruta alternativa tras avería
    if dato_averia is None or dato_averia["distancia"] == float('inf'):
        # Sin suministro
        _resaltar_camino(
            axes[1], G_averia, G_orig,
            [],
            f"AVERÍA {origen_f} → {destino_f} ({km_f} km)\n"
            f"⚠ {destino} queda SIN SUMINISTRO",
            color_camino="red",
            arista_cortada=(origen_f, destino_f)
        )
    else:
        _resaltar_camino(
            axes[1], G_averia, G_orig,
            dato_averia["camino"],
            f"AVERÍA {origen_f} → {destino_f} ({km_f} km)  →  Ruta alternativa\n"
            f"{dato_averia['distancia']:.0f} km  ·  {dato_averia['energia']:.1f} MWh  "
            f"({dato_normal['energia']-dato_averia['energia']:+.1f} MWh)",
            color_camino="red",
            arista_cortada=(origen_f, destino_f)
        )

    plt.suptitle(
        f"Distribución eléctrica: {ORIGEN} → {destino}  |  "
        f"Avería simulada en tramo {origen_f} → {destino_f}",
        fontsize=12, fontweight="bold"
    )
    plt.tight_layout()
    plt.show()


#   SIMULACIÓN EN TIEMPO REAL – ANIMACIÓN DE LA ENERGÍA VIAJANDO

def simular_tiempo_real(grafo, camino_normal, camino_averia, fallo, destino):
    """
    Muestra una animación en tiempo real donde la energía viaja salto a salto
    por la ruta óptima. Cuando llega al tramo averiado, la barra se detiene,
    aparece el aviso de avería y la energía continúa por la ruta alternativa.

    Diseñada para impresionar en una presentación en directo:
      - Barra de progreso que crece con cada salto
      - Nodos que se iluminan uno a uno
      - Pausa dramática en el punto de avería
      - Comparación final de ambas rutas en pantalla
    """
    import matplotlib.patches as mpatches
    import matplotlib.animation as animation
    import time

    G    = _nx(grafo)
    pos  = _pos()
    orig_f, dest_f, km_f = fallo

    # Calcular energías acumuladas salto a salto
    def energias_por_salto(camino):
        dist_acum = 0
        energias  = [ENERGIA_INICIAL]
        for i in range(len(camino) - 1):
            u, v = camino[i], camino[i+1]
            if G.has_edge(u, v):
                dist_acum += G[u][v]["weight"]
            energias.append(calcular_energia_final(dist_acum))
        return energias

    en_normal = energias_por_salto(camino_normal)
    en_averia = energias_por_salto(camino_averia) if camino_averia else []

    fig = plt.figure(figsize=(20, 13))
    fig.patch.set_facecolor("#0D1117")

    ax_mapa  = fig.add_axes([0.01, 0.28, 0.60, 0.68])   # mapa grande
    ax_barra = fig.add_axes([0.01, 0.12, 0.60, 0.10])   # barra de progreso
    ax_info  = fig.add_axes([0.63, 0.05, 0.35, 0.91])   # panel lateral

    for ax in [ax_mapa, ax_barra, ax_info]:
        ax.set_facecolor("#0D1117")

    ax_mapa.set_axis_off()
    ax_barra.set_axis_off()
    ax_info.set_axis_off()

    # Título global
    fig.text(0.5, 0.97,
             f"⚡  Simulación en tiempo real: {ORIGEN} → {destino}",
             ha="center", va="top", fontsize=15, fontweight="bold",
             color="white", fontfamily="monospace")

    def dibujar_fondo():
        ax_mapa.clear(); ax_mapa.set_axis_off()
        ax_mapa.set_facecolor("#0D1117")
        for ccaa, ciudades in CCAA.items():
            nodos = [c for c in ciudades if c in G.nodes()]
            if nodos:
                nx.draw_networkx_nodes(G, pos, nodelist=nodos,
                                       node_size=220,
                                       node_color=COLORES_CCAA[ccaa],
                                       alpha=0.2, ax=ax_mapa)
        nx.draw_networkx_edges(G, pos, arrowstyle="->", arrowsize=6,
                               edge_color="#222233", width=0.4, ax=ax_mapa)
        for n in G.nodes():
            ax_mapa.annotate(n, xy=pos[n], fontsize=4.5,
                             color="#444455", ha="center", va="center",
                             xytext=(0, 8), textcoords="offset points")

    def dibujar_panel(fase, nodo_actual, energia_actual,
                      camino_hasta_ahora, averia_activa=False):
        ax_info.clear(); ax_info.set_axis_off()
        ax_info.set_facecolor("#0D1117")

        y = 0.97
        def txt(texto, color="white", size=9, bold=False, dy=0.04):
            nonlocal y
            ax_info.text(0.05, y, texto, transform=ax_info.transAxes,
                         fontsize=size, color=color, fontweight="bold" if bold else "normal",
                         va="top", fontfamily="monospace")
            y -= dy

        txt("─" * 38, color="#334455", dy=0.025)
        txt("ESTADO DE LA RED", color="#58A6FF", size=11, bold=True)
        txt("─" * 38, color="#334455", dy=0.025)
        txt("")

        if averia_activa:
            txt("⚠  AVERÍA DETECTADA", color="#FF4444", size=10, bold=True)
            txt(f"   Tramo cortado:", color="#FF8888")
            txt(f"   {orig_f} → {dest_f}", color="#FF8888")
            txt(f"   ({km_f} km fuera de servicio)", color="#FF8888")
            txt("")
            txt("↪  Recalculando ruta...", color="#FFD700", size=9, bold=True)
            txt("")
        else:
            txt(f"Fase: {fase}", color="#58A6FF")
            txt("")

        txt(f"Nodo actual:", color="#AAAAAA", size=8)
        txt(f"  📍 {nodo_actual}", color="#FFD700", size=10, bold=True)
        txt("")
        txt(f"Energía transportada:", color="#AAAAAA", size=8)
        pct = energia_actual / ENERGIA_INICIAL * 100
        color_en = "#2ECC71" if pct > 70 else "#F39C12" if pct > 40 else "#E74C3C"
        txt(f"  ⚡ {energia_actual:.1f} MWh  ({pct:.1f}%)", color=color_en, size=10, bold=True)
        txt(f"  Pérdida: {ENERGIA_INICIAL - energia_actual:.1f} MWh", color="#888888", size=8)
        txt("")

        if camino_hasta_ahora:
            txt("Recorrido:", color="#AAAAAA", size=8)
            for i in range(len(camino_hasta_ahora)):
                icono = "🟢" if i == 0 else ("🔴" if i == len(camino_hasta_ahora)-1 else "🟡")
                txt(f"  {icono} {camino_hasta_ahora[i]}", color="#DDDDDD", size=8, dy=0.033)

    def dibujar_barra(progreso, color="#2ECC71", label=""):
        ax_barra.clear(); ax_barra.set_axis_off()
        ax_barra.set_facecolor("#0D1117")
        ax_barra.set_xlim(0, 1); ax_barra.set_ylim(0, 1)
        # Fondo barra
        ax_barra.add_patch(mpatches.FancyBboxPatch(
            (0.01, 0.2), 0.98, 0.5,
            boxstyle="round,pad=0.01", fc="#1C2333", ec="#334455", lw=1.5))
        # Relleno
        if progreso > 0:
            ax_barra.add_patch(mpatches.FancyBboxPatch(
                (0.01, 0.2), max(0.01, progreso * 0.98), 0.5,
                boxstyle="round,pad=0.01", fc=color, ec="none", alpha=0.85))
        ax_barra.text(0.5, 0.5, label, ha="center", va="center",
                      fontsize=9, color="white", fontweight="bold",
                      fontfamily="monospace", transform=ax_barra.transAxes)

    # Animación fase 1: ruta normal
    PAUSA_SALTO    = 0.9    # segundos entre saltos
    PAUSA_AVERIA   = 3.5    # pausa dramática en la avería
    PAUSA_REANUDA  = 0.7

    nodos_iluminados = []
    aristas_iluminadas = []
    aristas_averia_iluminadas = []

    def iluminar_nodo(nodo, color, size=500):
        nx.draw_networkx_nodes(G, pos, nodelist=[nodo],
                               node_size=size, node_color=color,
                               edgecolors="white", linewidths=1.5,
                               ax=ax_mapa)
        ax_mapa.annotate(nodo, xy=pos[nodo], fontsize=7,
                         color="white", ha="center", va="center",
                         fontweight="bold",
                         xytext=(0, 10), textcoords="offset points")

    def iluminar_arista(u, v, color, width=3):
        if G.has_edge(u, v):
            nx.draw_networkx_edges(G, pos, edgelist=[(u, v)],
                                   edge_color=color, width=width,
                                   arrowstyle="->", arrowsize=15, ax=ax_mapa)
            mid = ((pos[u][0]+pos[v][0])/2, (pos[u][1]+pos[v][1])/2)
            ax_mapa.annotate(f"{G[u][v]['weight']} km",
                             xy=mid, fontsize=7, color=color,
                             ha="center", fontweight="bold",
                             bbox=dict(boxstyle="round,pad=0.2",
                                       fc="#0D1117", ec=color, alpha=0.85))

    # Recorrido normal
    plt.ion(); plt.show()
    averia_idx = None
    for i, nodo in enumerate(camino_normal):
        dibujar_fondo()
        # Redibujar lo ya recorrido
        for j, n in enumerate(nodos_iluminados):
            color = "#2ECC71" if j == 0 else "#FFD700"
            iluminar_nodo(n, color, size=600 if j == 0 else 480)
        for (u, v) in aristas_iluminadas:
            iluminar_arista(u, v, "#27AE60", width=3)

        # Nodo actual
        color_nodo = "#2ECC71" if i == 0 else "#58A6FF"
        iluminar_nodo(nodo, color_nodo, size=700 if i == 0 else 550)
        if i > 0:
            iluminar_arista(camino_normal[i-1], nodo, "#27AE60", width=3.5)
            aristas_iluminadas.append((camino_normal[i-1], nodo))
        nodos_iluminados.append(nodo)

        # Comprobar si llegamos al tramo averiado
        if i < len(camino_normal) - 1:
            sig = camino_normal[i+1]
            if nodo == orig_f and sig == dest_f:
                averia_idx = i

        dibujar_panel(
            fase="Ruta óptima",
            nodo_actual=nodo,
            energia_actual=en_normal[i],
            camino_hasta_ahora=camino_normal[:i+1]
        )
        n_total = len(camino_normal) - 1
        progreso = i / n_total if n_total > 0 else 1
        dibujar_barra(progreso, "#27AE60",
                      f"Ruta óptima: {i}/{n_total} saltos  —  {en_normal[i]:.1f} MWh")
        plt.pause(PAUSA_SALTO)

        # Pausa dramática en la avería
        if averia_idx is not None and i == averia_idx:
            # Dibujar tramo cortado en rojo
            dibujar_panel(fase="⚠ AVERÍA", nodo_actual=nodo,
                          energia_actual=en_normal[i],
                          camino_hasta_ahora=camino_normal[:i+1],
                          averia_activa=True)
            if G.has_edge(orig_f, dest_f):
                nx.draw_networkx_edges(G, pos, edgelist=[(orig_f, dest_f)],
                                       edge_color="#FF4444", width=4,
                                       style="dashed",
                                       arrowstyle="->", arrowsize=18, ax=ax_mapa)
                mid = ((pos[orig_f][0]+pos[dest_f][0])/2,
                       (pos[orig_f][1]+pos[dest_f][1])/2)
                ax_mapa.annotate("✂ AVERÍA", xy=mid, fontsize=11,
                                 color="#FF4444", fontweight="bold", ha="center",
                                 bbox=dict(boxstyle="round,pad=0.4",
                                           fc="#0D1117", ec="#FF4444", alpha=0.95))
            dibujar_barra(progreso, "#E74C3C", "⚠  AVERÍA EN LA LÍNEA — Recalculando ruta...")
            plt.pause(PAUSA_AVERIA)
            break

    # Animación fase 2: ruta alternativa (si existe)
    if camino_averia and averia_idx is not None:
        # Punto de desvío: mismo inicio hasta el nodo antes del corte
        punto_desvio = averia_idx  # índice en camino_averia
        # Encontrar hasta dónde coincide camino_averia con camino_normal
        idx_averia_alt = 0
        for k, n in enumerate(camino_averia):
            if n == orig_f:
                idx_averia_alt = k
                break

        for i in range(idx_averia_alt, len(camino_averia)):
            nodo = camino_averia[i]
            dibujar_fondo()
            # Redibujar ruta normal hasta avería (tenue)
            for j in range(len(nodos_iluminados)):
                iluminar_nodo(nodos_iluminados[j],
                              "#1A5276" if j > 0 else "#2ECC71",
                              size=400 if j > 0 else 600)
            for (u, v) in aristas_iluminadas:
                iluminar_arista(u, v, "#1A5276", width=1.5)
            # Tramo cortado
            if G.has_edge(orig_f, dest_f):
                nx.draw_networkx_edges(G, pos, edgelist=[(orig_f, dest_f)],
                                       edge_color="#FF4444", width=3,
                                       style="dashed",
                                       arrowstyle="->", arrowsize=15, ax=ax_mapa)
            # Ruta alternativa recorrida hasta ahora
            for j in range(idx_averia_alt, i):
                iluminar_nodo(camino_averia[j], "#FFD700", size=480)
                if j > 0:
                    iluminar_arista(camino_averia[j-1], camino_averia[j],
                                    "#E67E22", width=3)

            iluminar_nodo(nodo, "#E74C3C" if i == len(camino_averia)-1 else "#FF8C00",
                          size=600)
            if i > idx_averia_alt:
                iluminar_arista(camino_averia[i-1], nodo, "#E67E22", width=3.5)

            dibujar_panel(
                fase="Ruta alternativa",
                nodo_actual=nodo,
                energia_actual=en_averia[i - idx_averia_alt],
                camino_hasta_ahora=camino_averia[:i+1],
                averia_activa=False
            )
            n_alt = len(camino_averia) - 1 - idx_averia_alt
            prog_alt = (i - idx_averia_alt) / n_alt if n_alt > 0 else 1
            dibujar_barra(prog_alt, "#E67E22",
                          f"Ruta alternativa: {i-idx_averia_alt}/{n_alt} saltos  —  "
                          f"{en_averia[i-idx_averia_alt]:.1f} MWh")
            plt.pause(PAUSA_REANUDA)

    # Pantalla final de comparación
    plt.ioff()
    dibujar_fondo()

    en_norm_final  = en_normal[-1]
    en_av_final    = en_averia[-1] if en_averia else 0
    dist_norm      = sum(G[camino_normal[i]][camino_normal[i+1]]["weight"]
                         for i in range(len(camino_normal)-1)
                         if G.has_edge(camino_normal[i], camino_normal[i+1]))
    dist_av        = (sum(G[camino_averia[i]][camino_averia[i+1]]["weight"]
                          for i in range(len(camino_averia)-1)
                          if G.has_edge(camino_averia[i], camino_averia[i+1]))
                      if camino_averia else 0)

    # Redibujar ambas rutas
    for i in range(len(camino_normal)-1):
        u, v = camino_normal[i], camino_normal[i+1]
        if (u, v) != (orig_f, dest_f):
            iluminar_arista(u, v, "#1A5276", width=1.5)
    if camino_averia:
        for i in range(len(camino_averia)-1):
            iluminar_arista(camino_averia[i], camino_averia[i+1], "#E67E22", width=2.5)
    if G.has_edge(orig_f, dest_f):
        nx.draw_networkx_edges(G, pos, edgelist=[(orig_f, dest_f)],
                               edge_color="#FF4444", width=3, style="dashed",
                               arrowstyle="->", arrowsize=15, ax=ax_mapa)

    iluminar_nodo(ORIGEN, "#2ECC71", 700)
    iluminar_nodo(destino, "#E74C3C", 700)

    # Panel comparación final
    ax_info.clear(); ax_info.set_axis_off()
    ax_info.set_facecolor("#0D1117")
    y = 0.97
    def txt2(texto, color="white", size=9, bold=False, dy=0.05):
        nonlocal y
        ax_info.text(0.05, y, texto, transform=ax_info.transAxes,
                     fontsize=size, color=color,
                     fontweight="bold" if bold else "normal",
                     va="top", fontfamily="monospace")
        y -= dy

    txt2("─" * 36, color="#334455", dy=0.025)
    txt2("RESULTADO FINAL", color="#58A6FF", size=12, bold=True)
    txt2("─" * 36, color="#334455", dy=0.025)
    txt2("")
    txt2("✅  Ruta óptima:", color="#2ECC71", size=10, bold=True)
    txt2(f"  {' → '.join(camino_normal)}", color="#AAFFAA", size=7.5, dy=0.04)
    txt2(f"  {dist_norm:.0f} km  |  {en_norm_final:.1f} MWh", color="#2ECC71")
    txt2("")
    if camino_averia:
        txt2("🔀  Ruta alternativa:", color="#E67E22", size=10, bold=True)
        txt2(f"  {' → '.join(camino_averia)}", color="#FFDDAA", size=7.5, dy=0.04)
        txt2(f"  {dist_av:.0f} km  |  {en_av_final:.1f} MWh", color="#E67E22")
        txt2("")
        txt2("📊  Impacto de la avería:", color="#58A6FF", size=10, bold=True)
        txt2(f"  +{dist_av-dist_norm:.0f} km de recorrido", color="#FFAAAA")
        txt2(f"  -{en_norm_final-en_av_final:.1f} MWh adicionales perdidos", color="#FFAAAA")
    else:
        txt2("⚠  Sin ruta alternativa:", color="#E74C3C", size=10, bold=True)
        txt2(f"  {destino} queda SIN SUMINISTRO", color="#FF6666")

    txt2("")
    txt2("✂  Avería simulada:", color="#FF4444", size=9, bold=True)
    txt2(f"  {orig_f} → {dest_f}  ({km_f} km)", color="#FF8888")

    dibujar_barra(1.0, "#2ECC71" if camino_averia else "#E74C3C",
                  "✅  Simulación completada")
    ax_mapa.set_title(
        f"Comparación final: ruta óptima (azul) vs. alternativa (naranja)",
        fontsize=9, color="white", fontweight="bold", pad=6)

    plt.tight_layout()
    plt.show()


def main():
    print(MENSAJE_CIERZO)

    grafo_original = construir_grafo()
    res_normal     = ejecutar_simulacion(grafo_original)
    dibujar_red_normal(grafo_original, res_normal)

    disponibles = sorted(
        r["ciudad"] for r in res_normal
        if r["distancia"] not in (0, float('inf'))
    )
    print("Provincias disponibles:")
    print(", ".join(disponibles))

    destino = input(
        "\n¿A qué provincia quieres enviar la energía desde Zaragoza? "
    ).strip()

    if destino not in grafo_original.adyacencia:
        print(f"\n  '{destino}' no está en el grafo. Comprueba el nombre exacto.")
        return

    # Análisis completo: ruta óptima + avería + mapas
    resultado = analizar_destino_con_fallo(grafo_original, destino)
    if resultado is None:
        return
    dato_normal, dato_averia, fallo = resultado

    # Exportar Excel junto al script, con manejo de archivo bloqueado
    ruta_script = os.path.dirname(os.path.abspath(__file__))
    ruta_excel  = os.path.join(ruta_script, "red_electrica.xlsx")
    grafo_fallo = construir_grafo()
    grafo_fallo.eliminar_conexion(fallo[0], fallo[1])
    res_fallo = ejecutar_simulacion(grafo_fallo)
    try:
        exportar_excel(res_normal, res_fallo, fallo, ruta_excel)
    except PermissionError:
        print(" No se pudo guardar el Excel: cierra el archivo si está abierto.")
        ruta_alt = os.path.join(ruta_script, "red_electrica_nuevo.xlsx")
        exportar_excel(res_normal, res_fallo, fallo, ruta_alt)

    # Simulación animada en tiempo real
    camino_averia = (dato_averia["camino"]
                     if dato_averia and dato_averia["distancia"] != float('inf')
                     else None)
    respuesta = input(
        "\n¿Quieres ver la simulación animada en tiempo real? (s/n): "
    ).strip().lower()
    if respuesta == "s":
        simular_tiempo_real(grafo_original,
                            dato_normal["camino"],
                            camino_averia,
                            fallo, destino)


if __name__ == "__main__":
    main()